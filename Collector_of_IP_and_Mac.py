from openpyxl import Workbook, load_workbook
import os
import socket
import getmac

FILE = "IP_List.xlsx"
User = ""
UserID = ""
Responsibility = ""
ResponsibilityID = ""
Building = ""
Room = ""
Phone = "053-950-"
IP = socket.gethostbyname(socket.getfqdn())
MAC = getmac.get_mac_address(interface=None, ip=IP, ip6=None, hostname=None, network_request=True).upper()
MAC = MAC.replace(':', '-')

def main():
    getUserInfo()
    printInfo()
    saveInfo()
    os.system('pause')

def saveInfo():
    if (os.path.isfile(FILE) == False):
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = "사용자"
        sheet['B1'] = "사용자 학번/교직원번호"
        sheet['C1'] = "책임자"
        sheet['D1'] = "책임자 교직원번호"
        sheet['E1'] = "장비분류"
        sheet['F1'] = "IP"
        sheet['G1'] = "MAC"
        sheet['H1'] = "건물명"
        sheet['I1'] = "호실"
        sheet['J1'] = "전화번호"
        workbook.save(FILE)
    workbook = load_workbook(FILE, read_only=False, data_only=True)
    sheet = workbook.active
    sheet.append([User, UserID, Responsibility, ResponsibilityID, "PC", IP, MAC, Building, Room, Phone])
    workbook.save(FILE)
    print("\nSaved Successfully!\n")

def printInfo():
    print("\n--------------- Input Status ------------------\n")
    print(f"사용자: {User} / {UserID}")
    print(f"책임자: {Responsibility} / {ResponsibilityID}")
    print(f"건물명: {Building} / {Room}")
    print(f"전화번호: {Phone}")
    print(f"IP: {IP} / Mac: {MAC}")
    print("\n----------------------------------------------")


def getUserInfo():
    global User
    global UserID
    global Responsibility
    global ResponsibilityID
    global Building
    global Room
    global Phone

    while not User:
        User = input("사용자명: ")
    while not UserID:
        UserID = input("사용자 ID: ")
    Responsibility = input("책임자명 / 사용자와 동일하면 아무것도 입력하지 말고 Enter: ")
    if (Responsibility == ""):
        Responsibility = User
        ResponsibilityID = UserID
    else: 
        while not ResponsibilityID:
            ResponsibilityID = input("책임자ID: ")
    while not Building:
        Building = input("건물명 (1 - E9 / 2 - IT4): ")
        if (Building == "1"):
            Building = "E9"
        elif (Building == "2"):
            Building = "IT4"
        else:
            Building = ""
    while not Room:
        Room = input("호실: ")
    tmp = input("전화번호 (053-950-xxxx) / 전화번호 대역이 다르면 아무것도 입력하지 말고 Enter: ")
    if not tmp:
        Phone = ""
        while not Phone:
            Phone = input("전화번호: ")
    else:
        Phone += tmp

main()