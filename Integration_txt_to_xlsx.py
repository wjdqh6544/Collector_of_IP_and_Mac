from openpyxl import Workbook, load_workbook
import os

Dest_FILE = "IP_List.xlsx"
Origin_FILE = "IP_List_for_Linux.txt"

def main():
    print()
    f = open(Origin_FILE, 'r')
    Length = len(f.readlines())
    f.seek(0, 0)
    for i in range(Length):
        tmp = f.readline().split(';')
        User = tmp[0]
        UserID = tmp[1]
        Responsibility = tmp[2]
        ResponsibilityID = tmp[3]
        IP = tmp[5]
        MAC = tmp[6]
        Building = tmp[7]
        Room = tmp[8]
        Phone = tmp[9]
        saveInfo(User, UserID, Responsibility, ResponsibilityID, IP, MAC, Building, Room, Phone)
        print("Integrating IP Information from .txt to .xlsx... (" + str(i+1) + "/" + str(Length) + ")")
    f.close()
    print("\nIntegrated Successfully!\n")

def saveInfo(User, UserID, Responsibility, ResponsibilityID, IP, MAC, Building, Room, Phone):
    if (os.path.isfile(Dest_FILE) == False):
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = "사용자"
        sheet['B1'] = "사용자 학번/교직원번호"
        sheet['C1'] = "책임자"
        sheet['D1'] = "책임자 교직원번호"
        sheet['E1'] = "장비분류"
        sheet['F1'] = "IP"
        sheet['G1'] = "MAC"
        sheet['H1'] = "건물명"
        sheet['I1'] = "호실"
        sheet['J1'] = "전화번호"
        workbook.save(Dest_FILE)
    workbook = load_workbook(Dest_FILE, read_only=False, data_only=True)
    sheet = workbook.active
    sheet.append([User, UserID, Responsibility, ResponsibilityID, "PC", IP, MAC, Building, Room, Phone])
    workbook.save(Dest_FILE)

main()